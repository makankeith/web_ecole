from django.shortcuts import render, redirect, get_object_or_404
from .models import Classe
from ECOLE.models import Ecole
from .forms import ClasseForm

from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Count, Q
from .models import Classe
from ECOLE.models import Ecole
from ECOLE.utils import get_annee_active  # Assure-toi que le chemin d'importation correspond à ton projet

def liste_classes(request, ecole_id):
    ecole = get_object_or_404(Ecole, id=ecole_id)
    annee_active = get_annee_active(ecole) # Récupération de la session/année active

    # On filtre les classes de l'école et on compte les inscriptions uniquement reliées à l'année en cours
    classes = Classe.objects.filter(ecole=ecole).select_related('classe_suivante').annotate(
        total_inscrits=Count(
            'inscription',
            filter=Q(inscription__annee_scolaire=annee_active)
        )
    )

    if request.method == "POST":
        nom = request.POST.get('nom')
        niveau = request.POST.get('niveau')
        frais = request.POST.get('frais', 0)
        mensualite = request.POST.get('mensualite', 0)
        
        Classe.objects.create(
            nom=nom, 
            niveau=niveau, 
            ecole=ecole, 
            frais_inscription=frais,
            mensualite=mensualite
        )
        return redirect('liste_classes', ecole_id=ecole.id)

    return render(request, 'classes/liste.html', {
        'ecole': ecole,
        'classes': classes,
        'annee_active': annee_active
    })

def ajouter_classe(request, ecole_id):
    ecole = get_object_or_404(Ecole, id=ecole_id)

    if request.method == "POST":
        # On passe l'école au formulaire pour le filtrage du dropdown
        form = ClasseForm(request.POST, ecole=ecole)
        if form.is_valid():
            classe = form.save(commit=False)
            classe.ecole = ecole # Association de l'école parente
            classe.save()
            return redirect('liste_classes', ecole_id=ecole.id)
    else:
        form = ClasseForm(ecole=ecole)

    return render(request, 'classes/ajouter.html', {'form': form, 'ecole': ecole})

def modifier_classe(request, ecole_id, classe_id):
    ecole = get_object_or_404(Ecole, id=ecole_id)
    classe = get_object_or_404(Classe, id=classe_id, ecole=ecole)
    
    if request.method == "POST":
        form = ClasseForm(request.POST, instance=classe, ecole=ecole)
        if form.is_valid():
            form.save()
            return redirect('liste_classes', ecole_id=ecole_id)
    else:
        form = ClasseForm(instance=classe, ecole=ecole)
    
    return render(request, 'classes/modifier.html', {
        'form': form,
        'classe': classe,
        'ecole': ecole
    })
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .models import Classe
from ECOLE.models import Ecole
from ECOLE.utils import get_annee_active
from ECOLE.models import Inscription 

def exporter_classe_excel(request, ecole_id, classe_id):
    ecole = get_object_or_404(Ecole, id=ecole_id)
    classe = get_object_or_404(Classe, id=classe_id, ecole=ecole)
    annee_active = get_annee_active(ecole)
    
    # 1. Création du classeur Excel en mémoire
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste Éléves"
    
    # Désactiver le quadrillage par défaut si tu veux (optionnel)
    ws.views.sheetView[0].showGridLines = True

    # 2. Styles réutilisables (Police, couleurs, bordures)
    font_titre = Font(name="Arial", size=16, bold=True, color="1E293B")
    font_annee = Font(name="Arial", size=11, italic=True, color="64748B")
    font_th = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_td = Font(name="Arial", size=10, color="1E293B")
    
    fill_th = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Bleu Indigo de ta charte
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Ligne alternée
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # 3. Écriture de l'En-tête personnalisé demandé
    ws.merge_cells('A2:E2')
    ws['A2'] = f"LISTE DES ÉLÈVES DE LA CLASSE : {classe.nom.upper()}"
    ws['A2'].font = font_titre
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A3:E3')
    annee_str = annee_active.nom if annee_active else "N/A"
    ws['A3'] = f"Année Scolaire : {annee_str} — Établissement : {ecole.nom}"
    ws['A3'].font = font_annee
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    # Ajustement de la hauteur des lignes d'en-tête
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 20

    # 4. Tableau : Les titres de colonnes (Ligne 5)
    headers = ["N°", "Nom", "Prénom(s)", "Mensualité (FCFA)", "Date Inscription"]
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header_title
        cell.font = font_th
        cell.fill = fill_th
        cell.alignment = Alignment(horizontal='center' if col_num in [1, 5] else 'left', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[5].height = 25

    # 5. Récupération et insertion des données réelles
    if annee_active:
        inscriptions = Inscription.objects.filter(
            classe=classe, 
            annee_scolaire=annee_active
        ).select_related('eleve').order_by('eleve__nom', 'eleve__prenom')
    else:
        inscriptions = []

    row_idx = 6
    for idx, ins in enumerate(inscriptions, 1):
        row_data = [
            idx,
            ins.eleve.nom.upper(),
            ins.eleve.prenom,
            ins.mensualite,
            ins.date_inscription.strftime('%d/%m/%Y')
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = font_td
            cell.border = thin_border
            
            # Alignements et formats spécifiques
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 4:
                cell.number_format = '#,##0' # Format monétaire propre sans décimales
                cell.alignment = Alignment(horizontal='right')
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal='center')
                
            # Effet Zèbre (lignes alternées)
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
                
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    # 6. Redimensionnement automatique des colonnes en fonction du texte
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 7. Préparation de la réponse HTTP pour le téléchargement
    filename = f"liste_{classe.nom.replace(' ', '_')}_{annee_str}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response